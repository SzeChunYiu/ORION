#!/usr/bin/env python3
"""Run the frozen P2 SWIFT V3 population-correction cross-review controller transport study.

The study is public-development evidence.  It preserves the earlier Zenodo V2
active-comparator adverse terminal and does not claim cold start, protected
confirmation, exact ASReview application execution, or population transport.
No source title or abstract is emitted to a result artifact.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import platform
import sys
import time
from itertools import combinations
from pathlib import Path
from typing import Any, Callable

import numpy as np
import openpyxl
import scipy
import sklearn
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import SGDClassifier
from sklearn.svm import LinearSVC


REVIEWS = ("BPA", "Fluoride", "Neuropain", "PFOS-PFOA", "Transgenerational")
PMID_REVIEWS = frozenset(("BPA", "PFOS-PFOA", "Transgenerational"))
WORKBOOK_BY_REVIEW = {
    "BPA": "source-audit/swift-ohat.xlsx",
    "Fluoride": "source-audit/swift-ohat.xlsx",
    "Neuropain": "source-audit/swift-camarades.xlsx",
    "PFOS-PFOA": "source-audit/swift-ohat.xlsx",
    "Transgenerational": "source-audit/swift-ohat.xlsx",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_json_sha256(value: Any) -> str:
    body = json.dumps(value, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(body).hexdigest()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def normalize_pmid(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    raw = str(value).strip()
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]
    return raw


def content_identity(title: str, abstract: str) -> str:
    body = normalize_text(f"{title} {abstract}").encode("utf-8")
    return hashlib.sha256(body).hexdigest()


def record_identity(review: str, content_id: str) -> str:
    return hashlib.sha256(f"{review}\0{content_id}".encode("utf-8")).hexdigest()


def load_pubmed_snapshot(path: Path) -> tuple[dict[str, tuple[str, str]], dict[str, Any]]:
    records: dict[str, tuple[str, str]] = {}
    duplicate_pmids = 0
    with path.open(encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            row = json.loads(line)
            if sorted(row) != ["abstract", "pmid", "title"]:
                raise ValueError(f"PubMed snapshot schema mismatch at line {line_number}")
            pmid = normalize_pmid(row["pmid"])
            value = (normalize_text(row["title"]), normalize_text(row["abstract"]))
            if pmid in records:
                duplicate_pmids += 1
                if records[pmid] != value:
                    raise ValueError(f"Conflicting PubMed snapshot rows for PMID {pmid}")
                continue
            records[pmid] = value
    receipt = {
        "unique_pmids": len(records),
        "duplicate_pmid_excess": duplicate_pmids,
        "pmid_key_set_sha256": hashlib.sha256("\n".join(sorted(records)).encode("ascii")).hexdigest(),
    }
    return records, receipt


def worksheet_rows(path: Path, sheet: str) -> tuple[list[str], Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Missing worksheet {sheet} in {path}")
    worksheet = workbook[sheet]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [normalize_text(value) for value in next(iterator)]
    return headers, (workbook, iterator)


def load_review(
    root: Path,
    review: str,
    protocol: dict[str, Any],
    pubmed: dict[str, tuple[str, str]],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook_path = root / WORKBOOK_BY_REVIEW[review]
    headers, (workbook, iterator) = worksheet_rows(workbook_path, review)
    positions = {name: index for index, name in enumerate(headers)}
    label_spec = protocol["labels"][review]
    required = {label_spec["column"]}
    if review in PMID_REVIEWS:
        required.add("PMID")
    else:
        required.update(("Title", "Abstract"))
    missing_columns = sorted(required - positions.keys())
    if missing_columns:
        workbook.close()
        raise ValueError(f"{review} missing columns: {missing_columns}")

    seen_content: set[str] = set()
    rows: list[dict[str, Any]] = []
    receipt: dict[str, Any] = {
        "source_rows": 0,
        "missing_pubmed_metadata": 0,
        "empty_text_rows": 0,
        "duplicate_content_excess": 0,
        "noncanonical_label_rows": 0,
    }
    try:
        for source_row_number, values in enumerate(iterator, start=2):
            receipt["source_rows"] += 1
            label = normalize_text(values[positions[label_spec["column"]]])
            if label == label_spec["positive"]:
                binary_label = 1
            elif label == label_spec["negative"]:
                binary_label = 0
            else:
                receipt["noncanonical_label_rows"] += 1
                continue

            if review in PMID_REVIEWS:
                pmid = normalize_pmid(values[positions["PMID"]])
                metadata = pubmed.get(pmid)
                if metadata is None:
                    receipt["missing_pubmed_metadata"] += 1
                    continue
                title, abstract = metadata
            else:
                title = normalize_text(values[positions["Title"]])
                abstract = normalize_text(values[positions["Abstract"]])

            if not title and not abstract:
                receipt["empty_text_rows"] += 1
                continue
            content_id = content_identity(title, abstract)
            if content_id in seen_content:
                receipt["duplicate_content_excess"] += 1
                continue
            seen_content.add(content_id)
            identity = record_identity(review, content_id)
            rows.append(
                {
                    "abstract": abstract,
                    "content_identity": content_id,
                    "label": binary_label,
                    "record_identity": identity,
                    "source_row_number": source_row_number,
                    "text": f"{title} {abstract}".strip(),
                    "title": title,
                }
            )
    finally:
        workbook.close()

    rows.sort(key=lambda row: row["record_identity"])
    labels = np.asarray([row["label"] for row in rows], dtype=np.int8)
    receipt.update(
        {
            "eligible_before_deduplication": len(rows) + receipt["duplicate_content_excess"],
            "canonical_rows": len(rows),
            "included_rows": int(labels.sum()),
            "excluded_rows": int(len(labels) - labels.sum()),
            "canonical_record_order_sha256": hashlib.sha256(
                "\n".join(row["record_identity"] for row in rows).encode("ascii")
            ).hexdigest(),
            "canonical_content_set_sha256": hashlib.sha256(
                "\n".join(sorted(row["content_identity"] for row in rows)).encode("ascii")
            ).hexdigest(),
        }
    )
    return rows, receipt


def initial_seed(review: str, labels: np.ndarray, identities: list[str]) -> list[int]:
    def key(index: int) -> str:
        return hashlib.sha256(f"{review}\0{identities[index]}".encode("ascii")).hexdigest()

    positives = np.flatnonzero(labels == 1)
    negatives = np.flatnonzero(labels == 0)
    if len(positives) == 0 or len(negatives) == 0:
        raise ValueError(f"{review} does not contain both seed classes")
    positive = min((int(index) for index in positives), key=key)
    negative = min((int(index) for index in negatives), key=key)
    return [positive, negative]


def complete_active_order(
    x: Any,
    labels: np.ndarray,
    seed: list[int],
    batch_size: int,
    model_factory: Callable[[], Any],
    score_function: Callable[[Any, Any], np.ndarray],
    weight_ratio: float | None = None,
) -> tuple[list[int], int]:
    selected = list(seed)
    remaining = np.ones(len(labels), dtype=bool)
    remaining[selected] = False
    fits = 0
    while remaining.any():
        model = model_factory()
        fit_kwargs: dict[str, Any] = {}
        if weight_ratio is not None:
            selected_labels = labels[selected]
            positives = int(selected_labels.sum())
            negatives = len(selected_labels) - positives
            if positives == 0 or negatives == 0:
                raise ValueError("Balanced comparator requires both observed classes")
            weights = np.where(selected_labels == 1, 1.0, positives / (weight_ratio * negatives))
            weights = weights * (len(weights) / float(weights.sum()))
            fit_kwargs["sample_weight"] = weights
        model.fit(x[selected], labels[selected], **fit_kwargs)
        fits += 1
        pool = np.flatnonzero(remaining)
        scores = np.asarray(score_function(model, x[pool]), dtype=float)
        ranked_pool = pool[np.argsort(-scores, kind="stable")]
        chosen = ranked_pool[:batch_size]
        selected.extend(int(index) for index in chosen)
        remaining[chosen] = False
    return selected, fits


def order_sha256(order: list[int], identities: list[str]) -> str:
    return hashlib.sha256("\n".join(identities[index] for index in order).encode("ascii")).hexdigest()


def order_metrics(order: list[int], labels: np.ndarray) -> dict[str, float]:
    if len(order) != len(labels) or len(set(order)) != len(labels):
        raise ValueError("Every arm must order every canonical row exactly once")
    total_positives = int(labels.sum())
    if total_positives == 0:
        raise ValueError("Recall metrics require at least one positive row")
    cumulative = np.cumsum(labels[order])

    def recall_at(fraction: float) -> float:
        screened = max(1, math.ceil(fraction * len(order)))
        return float(cumulative[screened - 1] / total_positives)

    target = math.ceil(0.95 * total_positives)
    effort_index = int(np.searchsorted(cumulative, target, side="left")) + 1
    effort = effort_index / len(order)
    return {
        "recall_at_005": recall_at(0.05),
        "recall_at_010": recall_at(0.10),
        "recall_at_020": recall_at(0.20),
        "fraction_screened_at_95_recall": effort,
        "wss_at_95": 0.95 - effort,
    }


def hash_text_list(values: list[str]) -> str:
    digest = hashlib.sha256()
    for value in values:
        body = value.encode("utf-8")
        digest.update(len(body).to_bytes(8, "big"))
        digest.update(body)
    return digest.hexdigest()


def check_file(root: Path, relative: str, expected: dict[str, Any]) -> dict[str, Any]:
    path = root / relative
    actual_bytes = path.stat().st_size if path.is_file() else None
    actual_sha256 = sha256_file(path) if path.is_file() else None
    return {
        "bytes_actual": actual_bytes,
        "bytes_expected": expected.get("bytes"),
        "passed": actual_bytes == expected.get("bytes") and actual_sha256 == expected.get("sha256"),
        "sha256_actual": actual_sha256,
        "sha256_expected": expected.get("sha256"),
    }


def binding_receipt(
    root: Path,
    source_freeze_path: Path,
    protocol_path: Path,
    implementation_path: Path,
) -> tuple[dict[str, Any], bool]:
    source_freeze = json.loads(source_freeze_path.read_text(encoding="utf-8"))
    implementation = json.loads(implementation_path.read_text(encoding="utf-8"))
    receipt: dict[str, Any] = {
        "freeze_files": {},
        "source_files": {},
    }
    freeze_files = {
        "source_and_rights": (source_freeze_path, implementation["source_and_rights_freeze_sha256"]),
        "protocol": (protocol_path, implementation["protocol_freeze_sha256"]),
        "runner": (Path(__file__).resolve(), implementation["runner_sha256"]),
    }
    if "population_correction" in implementation:
        correction = implementation["population_correction"]
        freeze_files["population_correction"] = (root / correction["path"], correction["sha256"])
    for name, (path, expected) in freeze_files.items():
        actual = sha256_file(path) if path.is_file() else None
        receipt["freeze_files"][name] = {"actual": actual, "expected": expected, "passed": actual == expected}

    source_expectations: dict[str, dict[str, Any]] = {}
    source_expectations.update(source_freeze["primary_source"]["files"])
    source_expectations[source_freeze["rights"]["article_html"]["path"]] = source_freeze["rights"]["article_html"]
    source_expectations[source_freeze["rights"]["loader_declaration"]["path"]] = source_freeze["rights"][
        "loader_declaration"
    ]
    source_expectations[source_freeze["provenance"]["csmed_commit_receipt"]["path"]] = source_freeze[
        "provenance"
    ]["csmed_commit_receipt"]
    source_expectations[source_freeze["provenance"]["csmed_code_license"]["path"]] = source_freeze[
        "provenance"
    ]["csmed_code_license"]
    source_expectations.update(source_freeze["provenance"]["metadata_only_audits"])
    source_expectations[source_freeze["pubmed_enrichment"]["snapshot"]["path"]] = source_freeze[
        "pubmed_enrichment"
    ]["snapshot"]
    source_expectations[source_freeze["pubmed_enrichment"]["snapshot_receipt"]["path"]] = source_freeze[
        "pubmed_enrichment"
    ]["snapshot_receipt"]
    for relative, expected in sorted(source_expectations.items()):
        receipt["source_files"][relative] = check_file(root, relative, expected)

    passed = all(item["passed"] for group in receipt.values() for item in group.values())
    receipt["passed"] = passed
    return receipt, passed


def population_ok(protocol: dict[str, Any], receipts: dict[str, dict[str, Any]], overlap: dict[str, Any]) -> bool:
    expected_before = protocol["population"]["expected_before_deduplication"]
    expected_duplicates = protocol["population"]["expected_duplicate_excess"]
    checks = []
    for review in REVIEWS:
        receipt = receipts[review]
        checks.extend(
            (
                receipt["eligible_before_deduplication"] == expected_before[review],
                receipt["duplicate_content_excess"] == expected_duplicates[review],
                receipt["noncanonical_label_rows"] == 0,
                receipt["empty_text_rows"] == 0,
                receipt["included_rows"] > 0,
                receipt["excluded_rows"] > 0,
            )
        )
    checks.extend(
        (
            sum(receipt["canonical_rows"] for receipt in receipts.values())
            == protocol["population"]["canonical_rows_expected"],
            overlap["max_pairwise_shared_content_identities"] == 68,
        )
    )
    return all(checks)


def execute(
    root: Path,
    source_freeze_path: Path,
    protocol_path: Path,
    implementation_path: Path,
    out_path: Path,
) -> None:
    start = time.monotonic()
    protocol = json.loads(protocol_path.read_text(encoding="utf-8"))
    implementation = json.loads(implementation_path.read_text(encoding="utf-8"))
    binding, binding_ok = binding_receipt(root, source_freeze_path, protocol_path, implementation_path)
    result: dict[str, Any] = {
        "binding_receipt": binding,
        "claim_scope": protocol["claim_scope"],
        "forbidden_claims": protocol["forbidden_claims"],
        "identity": protocol["identity"],
        "implementation_freeze_sha256": sha256_file(implementation_path),
        "preserved_terminal": protocol["preserved_terminal"],
        "protocol_freeze_sha256": sha256_file(protocol_path),
        "source_and_rights_freeze_sha256": sha256_file(source_freeze_path),
        "software": {
            "numpy": np.__version__,
            "openpyxl": openpyxl.__version__,
            "python": platform.python_version(),
            "scikit_learn": sklearn.__version__,
            "scipy": scipy.__version__,
        },
    }
    if not binding_ok:
        result.update(
            {
                "arms": {},
                "gates": {name: False for name in protocol["gates"]},
                "terminal": protocol["terminals"]["cannot_check"],
            }
        )
        out_path.write_text(json.dumps(result, sort_keys=True, indent=2) + "\n", encoding="utf-8")
        return

    print("binding_passed; loading frozen PubMed snapshot", flush=True)
    pubmed_path = root / "private-source/pubmed-snapshot.jsonl"
    pubmed, pubmed_receipt = load_pubmed_snapshot(pubmed_path)
    review_rows: dict[str, list[dict[str, Any]]] = {}
    population_receipts: dict[str, dict[str, Any]] = {}
    for review in REVIEWS:
        rows, receipt = load_review(root, review, protocol, pubmed)
        review_rows[review] = rows
        population_receipts[review] = receipt
        print(f"population_loaded review={review} canonical_rows={len(rows)}", flush=True)

    pairwise_overlap = []
    max_overlap = 0
    for review_a, review_b in combinations(REVIEWS, 2):
        a = {row["content_identity"] for row in review_rows[review_a]}
        b = {row["content_identity"] for row in review_rows[review_b]}
        count = len(a & b)
        max_overlap = max(max_overlap, count)
        pairwise_overlap.append(
            {"review_a": review_a, "review_b": review_b, "shared_content_identities": count}
        )
    overlap_receipt = {
        "max_pairwise_shared_content_identities": max_overlap,
        "pairwise": pairwise_overlap,
    }
    population_passed = population_ok(protocol, population_receipts, overlap_receipt)
    if not population_passed:
        result.update(
            {
                "arms": {},
                "gates": {
                    "G1_BINDING": True,
                    "G2_POPULATION": False,
                    **{name: False for name in protocol["gates"] if name not in ("G1_BINDING", "G2_POPULATION")},
                },
                "overlap_receipt": overlap_receipt,
                "population_receipts": population_receipts,
                "pubmed_receipt": pubmed_receipt,
                "terminal": protocol["terminals"]["cannot_check"],
            }
        )
        out_path.write_text(json.dumps(result, sort_keys=True, indent=2) + "\n", encoding="utf-8")
        return

    arms_by_review: dict[str, Any] = {}
    effects_by_review: dict[str, Any] = {}
    for review in REVIEWS:
        rows = review_rows[review]
        texts = [row["text"] for row in rows]
        identities = [row["record_identity"] for row in rows]
        labels = np.asarray([row["label"] for row in rows], dtype=np.int8)
        batch_size = max(10, math.ceil(0.002 * len(rows)))
        seed = initial_seed(review, labels, identities)
        seed_hash = hashlib.sha256("\n".join(identities[index] for index in seed).encode("ascii")).hexdigest()

        print(f"candidate_vectorize review={review} batch_size={batch_size}", flush=True)
        candidate_vectorizer = TfidfVectorizer(
            ngram_range=(1, 2),
            min_df=2,
            max_features=50000,
            sublinear_tf=True,
            lowercase=True,
        )
        x_candidate = candidate_vectorizer.fit_transform(texts)
        candidate_order, candidate_fits = complete_active_order(
            x=x_candidate,
            labels=labels,
            seed=seed,
            batch_size=batch_size,
            model_factory=lambda: SGDClassifier(
                loss="log_loss",
                class_weight="balanced",
                alpha=1e-5,
                max_iter=2000,
                tol=1e-4,
                random_state=20260823,
            ),
            score_function=lambda model, pool: model.predict_proba(pool)[:, 1],
        )
        print(f"comparator_vectorize review={review}", flush=True)
        comparator_vectorizer = TfidfVectorizer(
            ngram_range=(1, 2),
            min_df=1,
            max_df=0.95,
            sublinear_tf=True,
            lowercase=True,
        )
        x_comparator = comparator_vectorizer.fit_transform(texts)
        comparator_order, comparator_fits = complete_active_order(
            x=x_comparator,
            labels=labels,
            seed=seed,
            batch_size=batch_size,
            model_factory=lambda: LinearSVC(loss="squared_hinge", C=0.11, random_state=20260823),
            score_function=lambda model, pool: model.decision_function(pool),
            weight_ratio=9.8,
        )
        candidate_metrics = order_metrics(candidate_order, labels)
        comparator_metrics = order_metrics(comparator_order, labels)
        arms_by_review[review] = {
            "ASREVIEW_ELAS_U4_CADENCE_MATCHED_COMPONENTS": {
                "features": int(x_comparator.shape[1]),
                "metrics": comparator_metrics,
                "model_fits": comparator_fits,
                "order_sha256": order_sha256(comparator_order, identities),
            },
            "FIXED_LOGREG_CONTROLLER": {
                "features": int(x_candidate.shape[1]),
                "metrics": candidate_metrics,
                "model_fits": candidate_fits,
                "order_sha256": order_sha256(candidate_order, identities),
            },
            "adapter_receipt": {
                "batch_size": batch_size,
                "initial_seed_record_identities_sha256": seed_hash,
                "text_list_sha256": hash_text_list(texts),
            },
        }
        effects_by_review[review] = {
            metric: candidate_metrics[metric] - comparator_metrics[metric]
            for metric in candidate_metrics
        }
        print(f"review_complete review={review}", flush=True)

    effect_names = tuple(next(iter(effects_by_review.values())).keys())
    mean_effects = {
        metric: float(np.mean([effects_by_review[review][metric] for review in REVIEWS]))
        for metric in effect_names
    }
    primary_by_review = {
        review: effects_by_review[review]["recall_at_010"] for review in REVIEWS
    }
    worst_review = min(REVIEWS, key=lambda review: primary_by_review[review])
    candidate_wss = {
        review: arms_by_review[review]["FIXED_LOGREG_CONTROLLER"]["metrics"]["wss_at_95"]
        for review in REVIEWS
    }
    gates = {
        "G1_BINDING": True,
        "G2_POPULATION": True,
        "G3_PRIMARY_MARGIN": mean_effects["recall_at_010"] >= 0.05,
        "G4_WORK_SAVING": mean_effects["wss_at_95"] >= 0.0,
        "G5_HARM": primary_by_review[worst_review] >= -0.05,
        "G6_ABSOLUTE_WORK_SAVING": all(value > 0.0 for value in candidate_wss.values()),
    }
    terminal = protocol["terminals"]["positive"] if all(gates.values()) else protocol["terminals"]["negative"]
    result.update(
        {
            "arms_by_review": arms_by_review,
            "effects_by_review": effects_by_review,
            "elapsed_seconds": time.monotonic() - start,
            "failed_gates": [name for name, passed in gates.items() if not passed],
            "gates": gates,
            "mean_candidate_minus_comparator": mean_effects,
            "overlap_receipt": overlap_receipt,
            "population_receipts": population_receipts,
            "pubmed_receipt": pubmed_receipt,
            "terminal": terminal,
            "total_canonical_rows": sum(receipt["canonical_rows"] for receipt in population_receipts.values()),
            "worst_review_at_recall_010": {
                "candidate_minus_comparator": primary_by_review[worst_review],
                "review": worst_review,
            },
        }
    )
    result["result_payload_sha256"] = canonical_json_sha256(result)
    out_path.write_text(json.dumps(result, sort_keys=True, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--source-freeze", type=Path, required=True)
    parser.add_argument("--protocol", type=Path, required=True)
    parser.add_argument("--implementation", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    args = parser.parse_args()
    execute(
        root=args.root.resolve(),
        source_freeze_path=args.source_freeze.resolve(),
        protocol_path=args.protocol.resolve(),
        implementation_path=args.implementation.resolve(),
        out_path=args.out.resolve(),
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
