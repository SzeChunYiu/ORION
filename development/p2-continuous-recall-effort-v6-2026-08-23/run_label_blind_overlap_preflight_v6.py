#!/usr/bin/env python3
"""Label-blind population and overlap preflight for the P2 KIFMS V6 source.

Allowed KIFMS fields are stable key, title, abstract and PubMed identifier.
The three outcome fields are verified by header name only; their row values are
never indexed, retained, counted, hashed, emitted or used in a branch.

The SWIFT and V5 comparison bodies are used only to form identifier/content
sets.  No SWIFT or V5 labels or ranking outcomes are read.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


LABEL_HEADERS = ["noisy_inclusion", "expert_inclusion", "fulltext_inclusion"]
ALLOWED_HEADERS = ["key", "title", "abstract", "pubmed_id"]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def hash_lines(values: list[str]) -> str:
    return hashlib.sha256("\n".join(values).encode("ascii")).hexdigest()


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split())


def normalize_pmid(value: Any) -> str:
    text = normalize_text(value)
    match = re.search(r"(\d+)(?:/)?$", text)
    return match.group(1) if match else text


def content_identity(title: Any, abstract: Any) -> str:
    text = f"{normalize_text(title)} {normalize_text(abstract)}"
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def reconstruct_abstract(index: Any) -> str:
    if not isinstance(index, dict) or not index:
        return ""
    positions: list[tuple[int, str]] = []
    for token, offsets in index.items():
        if not isinstance(offsets, list):
            continue
        for offset in offsets:
            if isinstance(offset, int):
                positions.append((offset, str(token)))
    positions.sort()
    return " ".join(token for _, token in positions)


def load_expected_kifms(metadata_path: Path) -> dict[str, dict[str, Any]]:
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    expected = {}
    for name, receipt in metadata.items():
        attributes = receipt["file_metadata"]["data"]["attributes"]
        version = receipt["version_metadata"]["data"]
        expected[name] = {
            "file_id": receipt["file_metadata"]["data"]["id"],
            "guid": attributes["guid"],
            "version": attributes["current_version"],
            "bytes": attributes["size"],
            "sha256": attributes["extra"]["hashes"]["sha256"],
            "revision_download": version["links"]["download"],
        }
    return expected


def load_kifms(
    root: Path, expected: dict[str, dict[str, Any]]
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, Any]]:
    reviews: dict[str, list[dict[str, Any]]] = {}
    receipts: dict[str, Any] = {}
    for filename in sorted(expected):
        path = root / filename
        if not path.is_file():
            raise FileNotFoundError(path)
        actual_hash = sha256_file(path)
        if actual_hash != expected[filename]["sha256"]:
            raise ValueError(f"KIFMS hash mismatch: {filename}")
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle, delimiter=";")
            header = next(reader)
            if header[:3] != LABEL_HEADERS:
                raise ValueError(f"unexpected label header boundary in {filename}: {header[:3]}")
            index = {name: header.index(name) for name in ALLOWED_HEADERS}
            rows: list[dict[str, Any]] = []
            for row_number, row in enumerate(reader, start=2):
                # Do not index row[0], row[1] or row[2].
                title = row[index["title"]]
                abstract = row[index["abstract"]]
                pmid = normalize_pmid(row[index["pubmed_id"]])
                key = normalize_text(row[index["key"]]) or f"row-{row_number:09d}"
                rows.append(
                    {
                        "key": key,
                        "row_number": row_number,
                        "content_id": content_identity(title, abstract),
                        "empty_text": not (normalize_text(title) or normalize_text(abstract)),
                        "pmid": pmid,
                    }
                )
        review = filename.removesuffix(".csv")
        reviews[review] = rows
        receipts[review] = {
            **expected[filename],
            "filename": filename,
            "header": header,
            "label_headers_verified_only": LABEL_HEADERS,
            "label_values_accessed": False,
            "raw_rows": len(rows),
        }
    return reviews, receipts


def load_swift(ohat_path: Path, camrades_path: Path) -> dict[str, Any]:
    pmids: set[str] = set()
    content: set[str] = set()
    review_receipts: dict[str, Any] = {}
    for path, expected_sheets in [
        (ohat_path, ["BPA", "PFOS-PFOA", "Transgenerational", "Fluoride"]),
        (camrades_path, ["Neuropain"]),
    ]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in expected_sheets:
            worksheet = workbook[sheet]
            iterator = worksheet.iter_rows(values_only=True)
            header = [str(value or "") for value in next(iterator)]
            positions = {name: header.index(name) for name in header}
            rows = 0
            local_pmids: set[str] = set()
            local_content: set[str] = set()
            for row in iterator:
                rows += 1
                if sheet in {"BPA", "PFOS-PFOA", "Transgenerational"}:
                    pmid = normalize_pmid(row[positions["PMID"]])
                    if pmid:
                        local_pmids.add(pmid)
                else:
                    local_content.add(
                        content_identity(row[positions["Title"]], row[positions["Abstract"]])
                    )
            pmids.update(local_pmids)
            content.update(local_content)
            review_receipts[sheet] = {
                "raw_rows": rows,
                "pmid_identities": len(local_pmids),
                "content_identities": len(local_content),
                "labels_accessed": False,
            }
    return {
        "pmids": pmids,
        "content": content,
        "reviews": review_receipts,
        "source_files": {
            str(ohat_path): {"bytes": ohat_path.stat().st_size, "sha256": sha256_file(ohat_path)},
            str(camrades_path): {
                "bytes": camrades_path.stat().st_size,
                "sha256": sha256_file(camrades_path),
            },
        },
    }


def load_v5_works(root: Path, binding_path: Path) -> dict[str, Any]:
    binding = json.loads(binding_path.read_text(encoding="utf-8"))
    expected = {
        str(item["file_id"]): item
        for item in binding["files"]
        if item["filename"].startswith("works_")
    }
    pmids: set[str] = set()
    content: set[str] = set()
    work_ids: set[str] = set()
    archive_receipts = []
    for file_id in sorted(expected, key=int):
        path = root / f"{file_id}.zip"
        item = expected[file_id]
        actual_hash = sha256_file(path)
        if actual_hash != item["sha256"]:
            raise ValueError(f"V5 works hash mismatch: {file_id}")
        raw_rows = 0
        with zipfile.ZipFile(path) as archive:
            members = sorted(name for name in archive.namelist() if name.endswith(".json"))
            for member in members:
                works = json.loads(archive.read(member))
                if not isinstance(works, list):
                    raise ValueError(f"unexpected V5 works payload: {path}:{member}")
                for work in works:
                    raw_rows += 1
                    work_id = normalize_text(work.get("id"))
                    if work_id:
                        work_ids.add(work_id)
                    ids = work.get("ids") or {}
                    pmid = normalize_pmid(ids.get("pmid"))
                    if pmid:
                        pmids.add(pmid)
                    content.add(
                        content_identity(
                            work.get("title"), reconstruct_abstract(work.get("abstract_inverted_index"))
                        )
                    )
        archive_receipts.append(
            {
                "file_id": int(file_id),
                "review": item["review"],
                "bytes": path.stat().st_size,
                "sha256": actual_hash,
                "json_members": len(members),
                "raw_work_rows": raw_rows,
                "labels_accessed": False,
            }
        )
    return {
        "pmids": pmids,
        "content": content,
        "work_ids": work_ids,
        "archives": archive_receipts,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--kifms-root", type=Path, required=True)
    parser.add_argument("--kifms-metadata", type=Path, required=True)
    parser.add_argument("--swift-ohat", type=Path, required=True)
    parser.add_argument("--swift-camrades", type=Path, required=True)
    parser.add_argument("--v5-works-root", type=Path, required=True)
    parser.add_argument("--v5-binding", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    expected = load_expected_kifms(args.kifms_metadata)
    reviews, receipts = load_kifms(args.kifms_root, expected)
    swift = load_swift(args.swift_ohat, args.swift_camrades)
    v5 = load_v5_works(args.v5_works_root, args.v5_binding)

    pre_external: dict[str, list[dict[str, Any]]] = {}
    for review, source_rows in reviews.items():
        seen_content: set[str] = set()
        seen_pmids: set[str] = set()
        retained: list[dict[str, Any]] = []
        empty = duplicate_content = duplicate_pmid = 0
        for row in sorted(source_rows, key=lambda value: (value["key"], value["row_number"])):
            if row["empty_text"]:
                empty += 1
                continue
            if row["content_id"] in seen_content:
                duplicate_content += 1
                continue
            if row["pmid"] and row["pmid"] in seen_pmids:
                duplicate_pmid += 1
                continue
            seen_content.add(row["content_id"])
            if row["pmid"]:
                seen_pmids.add(row["pmid"])
            retained.append(row)
        swift_content = sum(row["content_id"] in swift["content"] for row in retained)
        swift_pmid = sum(bool(row["pmid"]) and row["pmid"] in swift["pmids"] for row in retained)
        v5_content = sum(row["content_id"] in v5["content"] for row in retained)
        v5_pmid = sum(bool(row["pmid"]) and row["pmid"] in v5["pmids"] for row in retained)
        external_free = [
            row
            for row in retained
            if row["content_id"] not in swift["content"]
            and (not row["pmid"] or row["pmid"] not in swift["pmids"])
            and row["content_id"] not in v5["content"]
            and (not row["pmid"] or row["pmid"] not in v5["pmids"])
        ]
        pre_external[review] = external_free
        receipts[review].update(
            {
                "empty_text_rows": empty,
                "within_review_duplicate_content_excess": duplicate_content,
                "within_review_duplicate_pmid_excess_after_content_dedup": duplicate_pmid,
                "provisional_unique_rows": len(retained),
                "provisional_nonempty_pmids": sum(bool(row["pmid"]) for row in retained),
                "swift_content_matches": swift_content,
                "swift_pmid_matches": swift_pmid,
                "v5_content_matches": v5_content,
                "v5_pmid_matches": v5_pmid,
                "external_disjoint_rows": len(external_free),
            }
        )

    content_reviews: dict[str, set[str]] = defaultdict(set)
    pmid_reviews: dict[str, set[str]] = defaultdict(set)
    for review, rows in pre_external.items():
        for row in rows:
            content_reviews[row["content_id"]].add(review)
            if row["pmid"]:
                pmid_reviews[row["pmid"]].add(review)
    shared_content = {value for value, owners in content_reviews.items() if len(owners) > 1}
    shared_pmids = {value for value, owners in pmid_reviews.items() if len(owners) > 1}

    final: dict[str, list[dict[str, Any]]] = {}
    for review, rows in pre_external.items():
        final_rows = [
            row
            for row in rows
            if row["content_id"] not in shared_content
            and (not row["pmid"] or row["pmid"] not in shared_pmids)
        ]
        final[review] = final_rows
        receipts[review].update(
            {
                "excluded_cross_review_content_rows": sum(
                    row["content_id"] in shared_content for row in rows
                ),
                "excluded_cross_review_pmid_rows": sum(
                    bool(row["pmid"]) and row["pmid"] in shared_pmids for row in rows
                ),
                "canonical_rows": len(final_rows),
                "canonical_nonempty_pmids": sum(bool(row["pmid"]) for row in final_rows),
                "canonical_content_set_sha256": hash_lines(
                    sorted(row["content_id"] for row in final_rows)
                ),
                "canonical_pmid_set_sha256": hash_lines(
                    sorted(row["pmid"] for row in final_rows if row["pmid"])
                ),
            }
        )

    canonical_content = sorted(
        {row["content_id"] for rows in final.values() for row in rows}
    )
    canonical_pmids = sorted({row["pmid"] for rows in final.values() for row in rows if row["pmid"]})
    result = {
        "identity": "P2_KIFMS_V6_LABEL_BLIND_SOURCE_AND_OVERLAP_PREFLIGHT",
        "source_family": "OSF vt3n4 Medical Guidelines Dutch Association Medical Specialists",
        "review_count": len(final),
        "review_units": sorted(final),
        "label_boundary": {
            "headers_verified_only": LABEL_HEADERS,
            "selected_future_label": "expert_inclusion",
            "label_values_accessed": False,
            "class_counts_accessed": False,
            "seeds_accessed": False,
            "rankings_or_model_outcomes_accessed": False,
        },
        "canonicalization": {
            "content_identity": "SHA-256(case-sensitive whitespace-normalized title + one ASCII space + abstract)",
            "pmid_identity": "terminal digit string from pubmed_id/URL",
            "within_review": "exclude empty text; stable ascending key then source row; retain first unique content and nonempty PMID",
            "external": "exclude on either content or PMID match to raw SWIFT or V5 work bodies",
            "cross_review": "exclude every row whose content identity or nonempty PMID occurs in more than one KIFMS review",
        },
        "external_references": {
            "swift": {
                "review_count": len(swift["reviews"]),
                "pmid_identities": len(swift["pmids"]),
                "content_identities": len(swift["content"]),
                "reviews": swift["reviews"],
                "source_files": swift["source_files"],
                "labels_accessed": False,
            },
            "synergy_v5_raw_works": {
                "archive_count": len(v5["archives"]),
                "work_identities": len(v5["work_ids"]),
                "pmid_identities": len(v5["pmids"]),
                "content_identities": len(v5["content"]),
                "archives": v5["archives"],
                "labels_downloaded_or_accessed": False,
            },
        },
        "overlap": {
            "raw_external_match_counts_are_nonexclusive": True,
            "shared_kifms_content_identities_before_exclusion": len(shared_content),
            "shared_kifms_pmid_identities_before_exclusion": len(shared_pmids),
            "final_swift_content_matches": sum(
                row["content_id"] in swift["content"] for rows in final.values() for row in rows
            ),
            "final_swift_pmid_matches": sum(
                bool(row["pmid"]) and row["pmid"] in swift["pmids"]
                for rows in final.values()
                for row in rows
            ),
            "final_v5_content_matches": sum(
                row["content_id"] in v5["content"] for rows in final.values() for row in rows
            ),
            "final_v5_pmid_matches": sum(
                bool(row["pmid"]) and row["pmid"] in v5["pmids"]
                for rows in final.values()
                for row in rows
            ),
        },
        "per_review": receipts,
        "total_raw_rows": sum(value["raw_rows"] for value in receipts.values()),
        "total_canonical_rows": sum(len(rows) for rows in final.values()),
        "total_canonical_nonempty_pmids": len(canonical_pmids),
        "canonical_union_content_set_sha256": hash_lines(canonical_content),
        "canonical_union_pmid_set_sha256": hash_lines(canonical_pmids),
        "status": "LABEL_BLIND_CONTENT_AND_PMID_DISJOINT_POPULATION_FROZEN; CLASS_EXISTENCE_AND_EXECUTION_CANNOT_CHECK",
    }
    args.output.write_text(json.dumps(result, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(
        f"reviews={len(final)} raw={result['total_raw_rows']} canonical={result['total_canonical_rows']} "
        f"shared_content={len(shared_content)} shared_pmids={len(shared_pmids)}"
    )


if __name__ == "__main__":
    main()
